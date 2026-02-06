"""Export cleaned data to a styled Excel workbook with Listings and Summary sheets."""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path("output/prague_apartments.xlsx")

# Colors — emerald & terracotta theme
HEADER_FILL = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
ROW_FILL_1 = PatternFill(start_color="D8F3DC", end_color="D8F3DC", fill_type="solid")
ROW_FILL_2 = PatternFill(start_color="B7E4C7", end_color="B7E4C7", fill_type="solid")
GREEN_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
RED_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(color="2D3436", size=10)
SUMMARY_HEADER_FONT = Font(bold=True, size=12, color="1B4332")
SUMMARY_FONT = Font(size=11)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="95D5B2"),
)


def export_excel(df: pd.DataFrame) -> Path:
    """Create styled Excel workbook with Listings and Summary sheets."""
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    _write_listings_sheet(wb, df)
    _write_summary_sheet(wb, df)

    wb.save(OUTPUT_PATH)
    print(f"  Saved Excel to {OUTPUT_PATH}")
    return OUTPUT_PATH


def _write_listings_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.active
    ws.title = "Listings"

    columns = [
        ("Title", "title", 40, None),
        ("Price (CZK)", "price_czk", 15, '#,##0 "CZK"'),
        ("Area", "area_sqm", 10, '#,##0 "m²"'),
        ("Rooms", "rooms", 8, None),
        ("District", "district", 20, None),
        ("Price/m²", "price_per_sqm", 14, '#,##0 "CZK"'),
        ("Labels", "labels", 30, None),
        ("Latitude", "lat", 12, "0.0000"),
        ("Longitude", "lon", 12, "0.0000"),
        ("URL", "url", 50, None),
    ]

    # Headers
    for col_idx, (header, _, width, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        fill = ROW_FILL_1 if row_idx % 2 == 0 else ROW_FILL_2
        for col_idx, (_, field, _, fmt) in enumerate(columns, 1):
            value = row.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            if fmt:
                cell.number_format = fmt

    # Auto-filter
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{len(df) + 1}"

    # Conditional formatting on price_per_sqm (column F)
    median_price = int(df["price_per_sqm"].median())
    price_range = f"F2:F{len(df) + 1}"
    ws.conditional_formatting.add(
        price_range,
        CellIsRule(
            operator="lessThan",
            formula=[str(median_price)],
            fill=GREEN_FILL,
            font=Font(color="006100"),
        ),
    )
    ws.conditional_formatting.add(
        price_range,
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=[str(median_price)],
            fill=RED_FILL,
            font=Font(color="9C0006"),
        ),
    )

    ws.freeze_panes = "A2"


def _write_summary_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Summary")

    # Overall stats
    ws.cell(row=1, column=1, value="Prague Apartments — Summary").font = Font(
        bold=True, size=14, color="1B4332"
    )
    ws.merge_cells("A1:D1")

    stats = [
        ("Total listings", len(df)),
        ("Average price", f"{df['price_czk'].mean():,.0f} CZK"),
        ("Median price", f"{df['price_czk'].median():,.0f} CZK"),
        ("Average price/m²", f"{df['price_per_sqm'].mean():,.0f} CZK"),
    ]
    for i, (label, value) in enumerate(stats, 3):
        ws.cell(row=i, column=1, value=label).font = SUMMARY_FONT
        ws.cell(row=i, column=2, value=value).font = Font(bold=True, size=11)

    # Top 10 districts
    row_offset = len(stats) + 5
    ws.cell(
        row=row_offset, column=1, value="Top 10 Districts by Listing Count"
    ).font = SUMMARY_HEADER_FONT
    row_offset += 1
    for col_idx, header in enumerate(["District", "Count", "Avg Price/m²"], 1):
        cell = ws.cell(row=row_offset, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    district_stats = (
        df.groupby("district")
        .agg(count=("district", "size"), avg_psm=("price_per_sqm", "mean"))
        .sort_values("count", ascending=False)
        .head(10)
    )
    for i, (district, row) in enumerate(district_stats.iterrows(), row_offset + 1):
        ws.cell(row=i, column=1, value=district).font = SUMMARY_FONT
        ws.cell(row=i, column=2, value=row["count"]).font = SUMMARY_FONT
        ws.cell(row=i, column=3, value=f"{row['avg_psm']:,.0f} CZK").font = SUMMARY_FONT

    # Room distribution
    room_offset = row_offset + 13
    ws.cell(
        row=room_offset, column=1, value="Room Distribution"
    ).font = SUMMARY_HEADER_FONT
    room_offset += 1
    for col_idx, header in enumerate(["Rooms", "Count", "Avg Price"], 1):
        cell = ws.cell(row=room_offset, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    room_stats = (
        df[df["rooms"] != ""]
        .groupby("rooms")
        .agg(count=("rooms", "size"), avg_price=("price_czk", "mean"))
        .sort_values("count", ascending=False)
    )
    for i, (rooms, row) in enumerate(room_stats.iterrows(), room_offset + 1):
        ws.cell(row=i, column=1, value=rooms).font = SUMMARY_FONT
        ws.cell(row=i, column=2, value=row["count"]).font = SUMMARY_FONT
        ws.cell(
            row=i, column=3, value=f"{row['avg_price']:,.0f} CZK"
        ).font = SUMMARY_FONT

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
